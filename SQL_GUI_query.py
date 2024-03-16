import sys
import sqlite3
from PyQt5.QtGui import QSyntaxHighlighter, QTextCharFormat, QFont, QTextCursor, QPixmap
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QTextEdit, QVBoxLayout, QWidget, QLabel, QSplitter, QFileDialog
from PyQt5.QtCore import Qt, QRegExp
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem, QVBoxLayout, QDialog
import openpyxl
from openpyxl.styles import Font

class SQLHighlighter(QSyntaxHighlighter):
    def __init__(self, parent=None):
        super(SQLHighlighter, self).__init__(parent)

        self.highlighting_rules = []

        keyword_format = QTextCharFormat()
        keyword_format.setForeground(Qt.darkBlue)
        keyword_format.setFontWeight(QFont.Bold)

        keywords = [
            "SELECT", "FROM", "WHERE", "AND", "OR", "ORDER BY", "GROUP BY", "HAVING", "INSERT",
            "UPDATE", "DELETE", "CREATE", "ALTER", "DROP", "TABLE", "VIEW", "INDEX", "JOIN",
            "INNER", "OUTER", "LEFT", "RIGHT", "FULL", "ON", "AS", "DISTINCT", "VALUES"
            # You can add more keywords, please use uppercase for new highlighting syntax, both cases will be processed below
        ]

        for word in keywords:
            pattern = QRegExp(r'\b' + word + r'\b', Qt.CaseInsensitive)
            rule = (pattern, keyword_format)
            self.highlighting_rules.append(rule)

    def highlightBlock(self, text):
        for pattern, format in self.highlighting_rules:
            expression = pattern
            index = expression.indexIn(text)
            while index >= 0:
                length = expression.matchedLength()
                self.setFormat(index, length, format)
                index = expression.indexIn(text, index + length)

class SQLQueryApp(QMainWindow):
    """SQL query application
    Args:
        QMainWindow (_type_): GUI main window
    """
    def __init__(self):
        super().__init__()

        self.setWindowTitle(self.tr("SQL Query App"))
        self.setGeometry(100, 100, 1200, 600)  # Set window size

        self.init_ui()

    def init_ui(self):
        background_label = QLabel(self)
        background_pixmap = QPixmap('GIO.jpg')  # Replace 'GIO.jpg' with the actual path to your image file
        background_label.setPixmap(background_pixmap)
        background_label.setGeometry(0, 0, background_pixmap.width(), background_pixmap.height())
        self.browse_button = QPushButton(self.tr("Browse Database"), self)
        self.browse_button.clicked.connect(self.browse_db_file)
        self.browse_button.setStyleSheet("background-color: #4CAF50; color: white;")
        
        self.db_path_label = QLabel(self.tr("Database File Path:"), self)
        self.db_path_label.setFixedHeight(30)  # Label height
        self.db_path_text_edit = QTextEdit(self)
        self.db_path_text_edit.setReadOnly(True)
        self.db_path_text_edit.setPlaceholderText(self.tr("Select the path to your database file"))
        self.db_path_text_edit.setFixedHeight(30)  # Text box height
        self.db_path_text_edit.setStyleSheet("font-size: 16px; color: red; background-color: black;")  # Font size, color and background color

        self.text_edit = QTextEdit(self)
        self.text_edit.setPlaceholderText(self.tr("Enter your SQL query here"))
        self.text_edit.setFixedHeight(100)
        self.text_edit.setStyleSheet("font-size: 16px;")  # Font size
        
        self.highlighter = SQLHighlighter(self.text_edit.document())

        self.execute_button = QPushButton(self.tr("Execute Query"), self)
        self.execute_button.clicked.connect(self.execute_query)
        self.execute_button.setStyleSheet("background-color: #4CAF50; color: white;")  # Button style

        self.connect_button = QPushButton(self.tr("Connect to Database"), self)
        self.connect_button.clicked.connect(self.connect_to_database)
        self.connect_button.setStyleSheet("background-color: #4CAF50; color: white;")  # Button style

        self.show_tables_button = QPushButton(self.tr("Show Tables and Schemas"), self)
        self.show_tables_button.clicked.connect(self.show_tables_and_schemas)
        self.show_tables_button.setStyleSheet("background-color: #4CAF50; color: white;")  # Button style

        self.result_text_edit = QTextEdit(self)
        self.result_text_edit.setFixedHeight(300)  # Adjust the height of the output text box
        self.result_text_edit.setStyleSheet("font-size: 14px; background-color: #F5F5F5;")  # Set the result text box style

        self.tables_and_schemas_text_edit = QTextEdit(self)
        self.tables_and_schemas_text_edit.setFixedHeight(300)  # Adjust the height of the output text box
        self.tables_and_schemas_text_edit.setStyleSheet("font-size: 14px; background-color: #F5F5F5;")  # Set the result text box style

        splitter = QSplitter()
        splitter.addWidget(self.result_text_edit)
        splitter.addWidget(self.tables_and_schemas_text_edit)

        layout = QVBoxLayout()
        layout.addWidget(self.browse_button)
        layout.addWidget(self.db_path_label)
        layout.addWidget(self.db_path_text_edit)
        layout.addWidget(self.db_path_label)
        layout.addWidget(self.db_path_text_edit)
        layout.addWidget(self.connect_button)
        layout.addWidget(self.text_edit)
        layout.addWidget(self.execute_button)
        layout.addWidget(self.show_tables_button)
        layout.addWidget(splitter)  # Use a splitter to display two text boxes

        layout.addWidget(self.connect_button)
        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

        self.conn = None  # Database connection object
    def browse_db_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_name, _ = QFileDialog.getOpenFileName(self, self.tr("Select Database File"), "", self.tr("SQLite Database Files (*.db *.sqlite *.db3);;All Files (*)"), options=options)

        if file_name:
            self.db_path_text_edit.setPlainText(file_name)
    def connect_to_database(self):
        db_path = self.db_path_text_edit.toPlainText()
        try:
            self.conn = sqlite3.connect(db_path)
            QMessageBox.information(self, self.tr("Connection Result"), self.tr("Connected to the database."))
        except sqlite3.Error as e:
            QMessageBox.critical(self, self.tr("Connection Error"), self.tr(f"Error: {e}"))

    def show_tables_and_schemas(self):
        if self.conn is None:
            self.result_text_edit.setPlainText(self.tr("Please connect to a database first."))
            return

        cursor = self.conn.cursor()

        # Query all tables
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()

        result_text = self.tr("Tables:\n")
        for table in tables:
            result_text += table[0] + "\n"

        # Query schema for each table
        for table in tables:
            table_name = table[0]
            cursor.execute(f"PRAGMA table_info([{table_name}]);")
            columns = cursor.fetchall()

            result_text += f"\n{self.tr('Schema for')} {table_name}:\n"
            for column in columns:
                result_text += f"{column[1]} {column[2]}\n"


        self.tables_and_schemas_text_edit.setPlainText(result_text)

        cursor.close()

    def execute_query(self):
        if self.conn is None:
            self.result_text_edit.setPlainText(self.tr("Please connect to a database first."))
            return

        query = self.text_edit.toPlainText()
        
        if not query:
            self.result_text_edit.setPlainText(self.tr("Please enter a valid SQL query."))
            return
        
        cursor = self.conn.cursor()

        try:
            cursor.execute(query)
            results = cursor.fetchall()
            num_columns = len(cursor.description)
            num_rows = len(results)
            

            table_dialog = QDialog(self)
            table_dialog.setWindowTitle(self.tr("Query Results"))

            table_widget = QTableWidget()
            table_widget.setRowCount(num_rows)
            table_widget.setColumnCount(num_columns)

            # Set table headers
            headers = [description[0] for description in cursor.description]
            table_widget.setHorizontalHeaderLabels(headers)

            for i, row in enumerate(results):
                for j, value in enumerate(row):
                    item = QTableWidgetItem(str(value))
                    table_widget.setItem(i, j, item)

            layout = QVBoxLayout()
            layout.addWidget(table_widget)

            save_button = QPushButton(self.tr("Save as Excel (.xlsx)"), table_dialog)
            save_button.clicked.connect(lambda: self.save_table_as_excel(table_widget, headers))
            layout.addWidget(save_button)

            table_dialog.setLayout(layout)

            # Add maximize button
            table_dialog.setWindowFlags(table_dialog.windowFlags() | Qt.WindowMaximizeButtonHint)

            table_dialog.exec_()

        except sqlite3.Error as e:
            self.result_text_edit.setPlainText(self.tr(f"Error: {e}"))

        cursor.close()
    def save_table_as_excel(self, table_widget, headers):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_name, _ = QFileDialog.getSaveFileName(self, self.tr("Save as Excel (.xlsx)"), "", self.tr("Excel Files (*.xlsx);;All Files (*)"), options=options)

        if file_name:
            if not file_name.endswith(".xlsx"):
                file_name += ".xlsx"

            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)

            # Write data
            for row_num in range(table_widget.rowCount()):
                for col_num in range(table_widget.columnCount()):
                    cell = sheet.cell(row=row_num + 2, column=col_num + 1)
                    cell.value = table_widget.item(row_num, col_num).text()

            workbook.save(file_name)
    
def main():
    app = QApplication(sys.argv)
    window = SQLQueryApp()

    # Style
    app.setStyle('Fusion')

    # Global font
    font = QFont()
    font.setPointSize(12)
    app.setFont(font)

    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()